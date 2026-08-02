import streamlit as st
import pdfplumber
import pandas as pd
import re, io, tempfile
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule

TORONTO_TZ = ZoneInfo("America/Toronto")

st.set_page_config(page_title="RBC → Excel", page_icon="🏦", layout="centered")
st.title("🏦 RBC Bank Statements → Excel")
st.caption("Upload one or more RBC PDFs — download all transactions as Excel")

MONTH_MAP = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
             "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
DATE = re.compile(r"^(\d{1,2})(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)", re.I)

# ── Categories: same rules as the formula written into Excel ──
CATEGORIES = ["Airbnb", "Transfer", "Wire Payment", "Bank Fee", "Cash Withdrawal", "Other"]

# Professional pastel palette — one color per category, shared by the Transactions
# sheet (Category column) and the Category Summary sheet (whole row).
CATEGORY_COLORS = {
    "Airbnb":          "FFF9D9DC",  # soft rose
    "Transfer":        "FFD9E6F7",  # soft sky blue
    "Wire Payment":    "FFE4DBF2",  # soft lavender
    "Bank Fee":        "FFFBE8D2",  # soft peach
    "Cash Withdrawal": "FFE6DDD2",  # soft taupe
    "Other":           "FFE4E7E9",  # soft gray
}
CATEGORY_TEXT_COLOR = "FF3A3A3A"  # dark charcoal, readable on every pastel above

CREDIT_COLOR = "FF1E7145"  # dark green — positive amounts (Credit / Credits)
DEBIT_COLOR  = "FFB22222"  # dark red   — negative amounts (Debit / Debits)


def get_years(text):
    pat = re.compile(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s*\d{1,2}[,\s]+(\d{4})", re.I)
    return [(MONTH_MAP[m.group(1).capitalize()[:3]], int(m.group(2))) for m in pat.finditer(text)]

def best_year(month, anchors):
    if not anchors: return datetime.now().year
    for pm, py in anchors:
        if pm == month: return py
    return min(anchors, key=lambda x: abs(x[0]-month))[1]

def to_num(s):
    s = s.strip()
    return float(s.replace(",","")) if re.match(r"^[\d,]+\.\d{2}$", s) else None

def parse_pdf(path):
    raw = []  # list of {date, desc, debit, credit, balance} — one per PDF row
    with pdfplumber.open(path) as pdf:
        anchors = get_years("\n".join(p.extract_text() or "" for p in pdf.pages))
        cur_date = None

        for page in pdf.pages:
            words = page.extract_words()
            if not words: continue

            # Detect column x-positions from header row
            x_desc = x_debit = x_credit = x_bal = None
            for w in words:
                t = w["text"]
                if t == "Description":  x_desc   = w["x0"]
                if "Cheques"   in t:    x_debit  = w["x0"]
                if "Deposits"  in t:    x_credit = w["x0"]
                if t == "Balance($)":   x_bal    = w["x0"]
            if not x_debit: continue

            # Find activity section start Y
            act_y = next((w["top"] for w in words if "AccountActivity" in w["text"].replace(" ","")), 0)

            # Group words by row Y
            by_y = defaultdict(list)
            for w in words:
                if w["top"] > act_y:
                    by_y[round(w["top"])].append(w)

            for y in sorted(by_y):
                rw = sorted(by_y[y], key=lambda w: w["x0"])
                line = "".join(w["text"] for w in rw)

                if re.search(r"Description|Cheques|Deposits|Balance\(\$\)|Closingbalance|AccountFees|\dof\d|Openingbalance", line, re.I):
                    continue

                # Assign words to columns by x position
                date_t = desc_t = debit_t = credit_t = bal_t = ""
                for w in rw:
                    x, t = w["x0"], w["text"]
                    if x < x_desc:                  date_t   += t + " "
                    elif x < x_debit:               desc_t   += t + " "
                    elif x < x_credit:              debit_t  += t + " "
                    elif x < x_bal:                 credit_t += t + " "
                    else:                           bal_t    += t + " "

                # Parse date if present
                dm = DATE.match(date_t.strip().replace(" ",""))
                if dm:
                    day = int(dm.group(1)); mon = MONTH_MAP[dm.group(2).capitalize()[:3]]
                    cur_date = datetime(best_year(mon, anchors), mon, day).date()

                if not cur_date: continue

                # Clean description: remove Airbnb ref numbers and hex hashes
                desc = re.sub(r"\b\d{9,15}\b", "", desc_t).strip()
                desc = re.sub(r"\b[0-9a-f]{30,}\b", "", desc).strip()

                # Skip rows that are purely reference numbers with no amounts
                if not desc and not debit_t.strip() and not credit_t.strip() and not bal_t.strip():
                    continue

                raw.append({
                    "date":    cur_date,
                    "desc":    desc,
                    "debit":   to_num(debit_t.strip()),
                    "credit":  to_num(credit_t.strip()),
                    "balance": to_num(bal_t.strip()),
                })

    # Merge: if a row has desc but no amounts, merge with the next row that has amounts
    merged = []
    i = 0
    while i < len(raw):
        r = raw[i]
        has_amt = r["debit"] or r["credit"] or r["balance"]
        if r["desc"] and not has_amt:
            # Look ahead for the amounts row
            if i+1 < len(raw) and raw[i+1]["date"] == r["date"] and not raw[i+1]["desc"]:
                nxt = raw[i+1]
                merged.append({**r, "debit": nxt["debit"], "credit": nxt["credit"], "balance": nxt["balance"]})
                i += 2
                continue
        merged.append(r)
        i += 1

    # Remove leftover orphan amount-only rows (already merged above)
    result = [r for r in merged if r["desc"]]

    return [{
        "Date":        r["date"],
        "Description": r["desc"],
        "Credit":      r["credit"],
        "Debit":       r["debit"],
        "Balance":     r["balance"],
    } for r in result]


# ── Styles (colors/fonts replicated from the reference template) ──────
# NOTE: all colors use an explicit "FF" alpha prefix (opaque). Without it,
# openpyxl can write fully-transparent colors that some Excel contexts
# (especially conditional-formatting fills) render as invisible.
NAVY      = "FF1F4E79"
BLUE      = "FF2E75B6"
LIGHT_BLU = "FFD6E3F8"
ROW_TINT  = "FFF2F7FC"
WHITE     = "FFFFFFFF"
GRID_GRAY = "FFD9D9D9"

THIN_GRAY   = Side(style="thin", color=GRID_GRAY)
MEDIUM_NAVY = Side(style="medium", color=NAVY)
DOUBLE_NAVY = Side(style="double", color=NAVY)

CELL_BORDER      = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
TOP_RULE_BORDER  = Border(left=THIN_GRAY, right=THIN_GRAY, top=MEDIUM_NAVY, bottom=THIN_GRAY)
HEADER_BORDER    = Border(top=MEDIUM_NAVY)
TOTAL_BORDER     = Border(top=DOUBLE_NAVY)


def style_title(ws, text, ncols, size=18):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name="Calibri", size=size, bold=True, color=NAVY)
    c.fill = PatternFill("solid", fgColor=LIGHT_BLU)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40 if size >= 18 else 35
    ws.row_dimensions[2].height = 6


def style_header(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = HEADER_BORDER
    ws.row_dimensions[row].height = 28


def style_data_row(ws, row, ncols, first_data_row, is_first, size=11):
    fill = ROW_TINT if is_first or row % 2 == first_data_row % 2 else WHITE
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(name="Calibri", size=size)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal=c.alignment.horizontal or "center", vertical="center")
        c.border = TOP_RULE_BORDER if row == first_data_row else CELL_BORDER
    ws.row_dimensions[row].height = 20 if size == 10 else 22


def style_total_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal=c.alignment.horizontal or "center", vertical="center")
        c.border = TOTAL_BORDER
    ws.row_dimensions[row].height = 25


def finalize_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 125


# ── Build the final Excel workbook (3 sheets) ───────────────────────────
def build_excel(df: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    # ---------- Sheet 1: Transactions ----------
    ws = wb.create_sheet("Transactions")
    headers = ["Year", "Month", "Day", "Date", "Description",
               "Credit", "Debit", "Balance (PDF)", "Balance (Formula)", "File", "Category"]
    ncols = len(headers)

    style_title(ws, "📊 Bank Account Statement - Transactions", ncols, size=18)
    style_header(ws, 3, headers)

    first_data_row = 4
    n = len(df)
    last_data_row = first_data_row + n - 1

    for i, row in enumerate(df.itertuples(index=False)):
        r = first_data_row + i
        ws.cell(row=r, column=1, value=row.Date.year)
        ws.cell(row=r, column=2, value=row.Date.month)
        ws.cell(row=r, column=3, value=row.Date.day)
        dcell = ws.cell(row=r, column=4, value=datetime.combine(row.Date, datetime.min.time()))
        dcell.number_format = "mm-dd-yy"
        ws.cell(row=r, column=5, value=row.Description)
        if pd.notna(row.Credit):
            ws.cell(row=r, column=6, value=row.Credit)
        if pd.notna(row.Debit):
            ws.cell(row=r, column=7, value=row.Debit)
        # Balance (PDF): the real value extracted from the bank's statement (not a
        # formula, since it's the bank's official balance, not something derived here)
        if pd.notna(row.Balance):
            ws.cell(row=r, column=8, value=row.Balance)
        # Balance (Formula): recomputed from Credit/Debit. Row 1 is anchored to the
        # real PDF balance, and every row after that chains forward with a formula.
        if i == 0:
            ws.cell(row=r, column=9, value=row.Balance)
        else:
            ws.cell(row=r, column=9, value=f"=I{r-1}+F{r}-G{r}")
        ws.cell(row=r, column=10, value=row.File)
        # Category: dynamic formula based on the description (recalculates if edited)
        ws.cell(row=r, column=11, value=(
            f'=IF(ISNUMBER(SEARCH("AIRBNB",E{r})),"Airbnb",'
            f'IF(ISNUMBER(SEARCH("transfer",E{r})),"Transfer",'
            f'IF(ISNUMBER(SEARCH("wire",E{r})),"Wire Payment",'
            f'IF(ISNUMBER(SEARCH("fee",E{r})),"Bank Fee",'
            f'IF(ISNUMBER(SEARCH("withdrawal",E{r})),"Cash Withdrawal","Other")))))'
        ))

        for col in (6, 7, 8, 9):
            ws.cell(row=r, column=col).number_format = '"$"#,##0.00'

        style_data_row(ws, r, ncols, first_data_row, is_first=(i == 0), size=10)

        # Positive amounts (Credit) in dark green, negative amounts (Debit) in dark
        # red. Balance columns are left untouched, as requested.
        if pd.notna(row.Credit):
            ws.cell(row=r, column=6).font = Font(name="Calibri", size=10, color=CREDIT_COLOR)
        if pd.notna(row.Debit):
            ws.cell(row=r, column=7).font = Font(name="Calibri", size=10, color=DEBIT_COLOR)

    total_row = last_data_row + 1
    ws.cell(row=total_row, column=5, value="TOTAL")
    ws.cell(row=total_row, column=6, value=f"=SUM(F{first_data_row}:F{last_data_row})")
    ws.cell(row=total_row, column=7, value=f"=SUM(G{first_data_row}:G{last_data_row})")
    ws.cell(row=total_row, column=8, value=f"=F{total_row}-G{total_row}")
    # Balance (Formula) is not totaled: it's a running balance, not something summable
    for col in (6, 7, 8):
        ws.cell(row=total_row, column=col).number_format = '"$"#,##0.00'
    style_total_row(ws, total_row, ncols)

    widths = {"A": 8.33, "B": 6.66, "C": 6.66, "D": 15, "E": 46.66,
              "F": 14.16, "G": 14.16, "H": 15, "I": 15, "J": 41.66, "K": 16.66}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Category colors: conditional formatting on column K, so it stays in sync
    # automatically if the user edits a description and the category changes.
    # Note: for conditional-formatting fills, Excel needs both start_color and
    # end_color set (fgColor-only fills can render invisible in real Excel).
    for cat, color in CATEGORY_COLORS.items():
        ws.conditional_formatting.add(
            f"K{first_data_row}:K{last_data_row}",
            FormulaRule(
                formula=[f'$K{first_data_row}="{cat}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
                font=Font(name="Calibri", size=10, bold=True, color=CATEGORY_TEXT_COLOR),
                stopIfTrue=True,
            ),
        )

    ws.freeze_panes = "A4"
    finalize_sheet(ws)

    # ---------- Sheet 2: Monthly Summary ----------
    ws2 = wb.create_sheet("Monthly Summary")
    headers2 = ["Month", "Credits", "Debits", "Net"]
    style_title(ws2, "📅 Monthly Summary", len(headers2), size=16)
    style_header(ws2, 3, headers2)

    months = sorted({(d.year, d.month) for d in df["Date"]})
    m_first = 4
    for i, (y, m) in enumerate(months):
        r = m_first + i
        ws2.cell(row=r, column=1, value=f"{y}-{m:02d}")
        ws2.cell(row=r, column=2, value=(
            f"=SUMPRODUCT((YEAR(Transactions!$D${first_data_row}:$D${last_data_row})={y})*"
            f"(MONTH(Transactions!$D${first_data_row}:$D${last_data_row})={m})*"
            f"(Transactions!$F${first_data_row}:$F${last_data_row}))"
        ))
        ws2.cell(row=r, column=3, value=(
            f"=SUMPRODUCT((YEAR(Transactions!$D${first_data_row}:$D${last_data_row})={y})*"
            f"(MONTH(Transactions!$D${first_data_row}:$D${last_data_row})={m})*"
            f"(Transactions!$G${first_data_row}:$G${last_data_row}))"
        ))
        ws2.cell(row=r, column=4, value=f"=B{r}-C{r}")
        for col in (2, 3, 4):
            ws2.cell(row=r, column=col).number_format = '"$"#,##0.00'
        style_data_row(ws2, r, len(headers2), m_first, is_first=(i == 0), size=11)
        ws2.cell(row=r, column=2).font = Font(name="Calibri", size=11, color=CREDIT_COLOR)
        ws2.cell(row=r, column=3).font = Font(name="Calibri", size=11, color=DEBIT_COLOR)

    m_last = m_first + len(months) - 1
    m_total = m_last + 1
    ws2.cell(row=m_total, column=1, value="TOTAL")
    ws2.cell(row=m_total, column=2, value=f"=SUM(B{m_first}:B{m_last})")
    ws2.cell(row=m_total, column=3, value=f"=SUM(C{m_first}:C{m_last})")
    ws2.cell(row=m_total, column=4, value=f"=SUM(D{m_first}:D{m_last})")
    for col in (2, 3, 4):
        ws2.cell(row=m_total, column=col).number_format = '"$"#,##0.00'
    style_total_row(ws2, m_total, len(headers2))

    ws2.column_dimensions["A"].width = 16.66
    ws2.column_dimensions["B"].width = 18.33
    ws2.column_dimensions["C"].width = 18.33
    ws2.column_dimensions["D"].width = 16.66
    ws2.freeze_panes = "A4"
    finalize_sheet(ws2)

    # ---------- Sheet 3: Category Summary ----------
    ws3 = wb.create_sheet("Category Summary")
    headers3 = ["Category", "Count", "Percentage"]
    style_title(ws3, "🏷️ Category Summary", len(headers3), size=16)
    style_header(ws3, 3, headers3)

    c_first = 4
    for i, cat in enumerate(CATEGORIES):
        r = c_first + i
        ws3.cell(row=r, column=1, value=cat)
        ws3.cell(row=r, column=2, value=f"=COUNTIF(Transactions!$K:$K,A{r})")
        c_last_preview = c_first + len(CATEGORIES) - 1
        ws3.cell(row=r, column=3, value=f"=B{r}/SUM($B${c_first}:$B${c_last_preview})")
        ws3.cell(row=r, column=3).number_format = "0.0%"
        style_data_row(ws3, r, len(headers3), c_first, is_first=(i == 0), size=11)

        # Color the whole row with this category's color (same palette as Transactions)
        color = CATEGORY_COLORS.get(cat, "E4E7E9")
        for col in (1, 2, 3):
            cc = ws3.cell(row=r, column=col)
            cc.fill = PatternFill("solid", fgColor=color)
            cc.font = Font(name="Calibri", size=11, bold=True, color=CATEGORY_TEXT_COLOR)
        ws3.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")

    c_last = c_first + len(CATEGORIES) - 1
    c_total = c_last + 1
    ws3.cell(row=c_total, column=1, value="TOTAL")
    ws3.cell(row=c_total, column=2, value=f"=SUM(B{c_first}:B{c_last})")
    ws3.cell(row=c_total, column=3, value=f"=SUM(C{c_first}:C{c_last})")
    ws3.cell(row=c_total, column=3).number_format = "0.0%"
    style_total_row(ws3, c_total, len(headers3))

    ws3.column_dimensions["A"].width = 21.66
    ws3.column_dimensions["B"].width = 16.66
    ws3.column_dimensions["C"].width = 16.66
    ws3.freeze_panes = "A4"
    finalize_sheet(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── UI ─────────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "Upload RBC PDFs", type="pdf", accept_multiple_files=True,
    help="You can upload several months at once"
)

if uploaded:
    all_rows = []
    for f in uploaded:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(f.read())
        try:
            rows = parse_pdf(Path(tmp.name))
            for r in rows: r["File"] = f.name
            all_rows.extend(rows)
            st.success(f"✅ {f.name} → {len(rows)} transactions")
        except Exception as e:
            st.error(f"❌ {f.name}: {e}")

    if all_rows:
        df = pd.DataFrame(all_rows).reset_index(drop=True)
        df["Date"] = pd.to_datetime(df["Date"]).dt.date
        df = df.sort_values("Date").reset_index(drop=True)

        preview = df.copy()
        preview["Year"] = preview["Date"].apply(lambda d: d.year)
        preview["Month"] = preview["Date"].apply(lambda d: d.month)
        preview["Day"] = preview["Date"].apply(lambda d: d.day)
        preview = preview[["Year","Month","Day","Date","Description","Credit","Debit","Balance","File"]]

        st.divider()
        st.subheader(f"📋 {len(df)} transactions")
        st.dataframe(preview, use_container_width=True, hide_index=True)

        buf = build_excel(df)

        timestamp = datetime.now(TORONTO_TZ).strftime("%Y-%m-%d_%H-%M")
        filename = f"RBC_Analysis_Bank_Statements_{timestamp}.xlsx"

        st.download_button(
            "⬇️ Download Excel",
            data=buf, file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=True,
        )
